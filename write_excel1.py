from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws["A1"] = "こんにちは、takahiro!"
ws["B1"] = "Excelとの友情再開だね！"
wb.save("output1.xlsx")
