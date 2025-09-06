from openpyxl import load_workbook

wb = load_workbook("logbook.xlsx")
ws = wb["Log"]

#------1行目を全部表示
for row in ws.iter_rows(min_row=1,max_row=1,values_only=True):
    print(row)

#------3列目だけを全部取り出してリストにする
messages = []
for row in ws.iter_rows(min_row=2,values_only=True):
    messages.append(row[2])

print(messages)

#------空じゃないカテゴリだけ出力
for row in ws.iter_rows(min_row=2,values_only=True):
    if row[1]:
        print(row[1])

#------カテゴリが"業務”の行だけ内容を表示
for row in ws.iter_rows(min_row=2,values_only=True):
    if row[1] == "業務":
        print(row[2])

#------日時と内容をセットで表示
for row in ws.iter_rows(min_row=2,values_only=True):
    print(f"{row[0]}:{row[2]}")

#------内容に"git"が含まれる行だけ表示
for row in ws.iter_rows(min_row=2,values_only=True):
    if row[2]and "Git" in row[2]
        print(row[2])

#------カテゴリの種類を全部表示
categories = set()
for row in ws.iter_rows(min_row=2,values_only=True):
    if row[1]:
        categories.add([row[1]])

print(categories)