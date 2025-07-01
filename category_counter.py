from openpyxl import load_workbook
from datetime import datetime
from collections import Counter

filename = "logbook.xlsx"
wb = load_workbook(filename)
ws = wb["Log"]

#入力を受け取る
category = input("カテゴリを入力(学習/業務/休憩/日記など)：")
message = input("ログ内容を入力：")
now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

#ログを追記
ws.append([now, category, message])
wb.save(filename)
print("ログ記録を完了")

#カテゴリ別件数のカウント
categories = [row[1].value for row in ws.iter_rows(min_row=2) if row[1].value]
counted = Counter(categories)

print("\n カテゴリ別ログ件数")
for cat,cnt in counted.items():
    print(f"/{cat}:{cnt}回")
