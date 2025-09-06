from openpyxl import load_workbook, Workbook
from datetime import datetime
import os

# 元のExcelファイル
src_file = "logbook.xlsx"
try:
    wb_src = load_workbook(src_file)
    ws_src = wb_src["Log"]
except:
    print("元データが読み込めませんでした💥")
    exit()

# グループ用辞書: {(カテゴリ, 年): [データ]}
groups = {}

for row in ws_src.iter_rows(min_row=2, values_only=True):
    date_str, category, content = row

    # データが不足していたらスキップ
    if not (date_str and category and content):
        continue

    try:
        year = datetime.strptime(str(date_str), "%Y-%m-%d").year
    except:
        print(f"日付形式エラー: {date_str}")
        continue

    key = (category, year)
    if key not in groups:
        groups[key] = []
    groups[key].append((date_str, content))

# カテゴリと年別でExcelファイルを作成
for (category, year), items in groups.items():
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = "Log"
    ws_new.append(["日付", "内容"])

    for date, content in items:
        ws_new.append([date, content])

    # ファイル名に注意（全角や空白を避ける）
    safe_filename = f"{category}_{year}.xlsx".replace(" ", "_")
    wb_new.save(safe_filename)
    print(f"✅ {safe_filename} を出力しました！")
