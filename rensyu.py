from openpyxl import load_workbook,Workbook
from datetime import datetime
import os

src_file = load_workbook("logbook.xlsx")
try:
    wb_src = src_file
    ws_src = wb_src["Log"]
except:
    print("データを読み込めませんでした")
    exit

groups = {}

for row in ws_src.iter_rows(min_row=2,values_only=True):
    date_str,category,content = row
    if not(date_str and category and content):
        continue

    try:
        year = datetime.strptime(str(date_str),"%Y-%m-%d").year
    except:
        print("日時形式エラー")
        continue

    key = (category,year)
    if key not in groups:
        groups[key] = []
        groups[key].append(date_str,content)

for (category,year), items in groups.items():
    wb_new = Workbook()
    ws_new = wb_new.active
    wb_new.title = "Log"
    wb_new.append(["日付","内容"])

    for date,content in items:
        ws_new.append([date,content])

safe_filename = f"{category}_{year}.xlsx".replace("","_")
wb_new.save(safe_filename)
print(f"{safe_filename}を保存しました")