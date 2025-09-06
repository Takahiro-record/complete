from openpyxl import load_workbook,Workbook

years = range(2020,2025)#2025の一つ前までが結果として出る
output_wb = Workbook()
output_ws = output_wb.active
output_ws.title = "まとめ"
output_ws.append(["年","番号","内容","日付"])#最初にヘッダーを入力しておく

for year in years:
    fname = f"{year}.xlsx"
    try:
        wb = load_workbook(fname)
        ws = wb["Log"]
    except:
        print(f"{fname}読み込み失敗、スキップ")
        continue

    row_num = 1
    for row in ws.iter_rows(min_row=2,values_only=True):
        date,category,content = row
        if category == "学習" and content:
            output_ws.append([year,row_num,content,date])
            row_num += 1
        
output_wb.save("matome.xlsx")
print("まとめファイル作成完了")